#!/usr/bin/env python3
"""Creates the initial cars.xlsx template with all 50 brands and sample data."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- Styles ----
header_font  = Font(bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", fgColor="1A1A1A")
red_fill     = PatternFill("solid", fgColor="E8001C")
alt_fill     = PatternFill("solid", fgColor="F5F5F5")
center_align = Alignment(horizontal="center", vertical="center")
thin_border  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_header(ws, cols):
    ws.row_dimensions[1].height = 22
    for col, title in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = red_fill
        c.alignment = center_align
        c.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = max(len(title) + 6, 18)

def style_rows(ws, num_rows):
    for row in range(2, num_rows + 2):
        fill = alt_fill if row % 2 == 0 else PatternFill()
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            if fill.fill_type:
                c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(vertical="center")

# ================================================================
# Sheet 1: BRANDS
# ================================================================
ws_brands = wb.active
ws_brands.title = "Brands"
cols_brands = ["brand_id", "brand_name", "logo_file"]
style_header(ws_brands, cols_brands)

brands = [
    ("audi",        "Audi",        "audi.png"),
    ("bmw",         "BMW",         "bmw.png"),
    ("mercedes",    "Mercedes",    "mercedes.png"),
    ("volkswagen",  "Volkswagen",  "volkswagen.png"),
    ("ford",        "Ford",        "ford.png"),
    ("opel",        "Opel",        "opel.png"),
    ("peugeot",     "Peugeot",     "peugeot.png"),
    ("renault",     "Renault",     "renault.png"),
    ("toyota",      "Toyota",      "toyota.png"),
    ("honda",       "Honda",       "honda.png"),
    ("mazda",       "Mazda",       "mazda.png"),
    ("nissan",      "Nissan",      "nissan.png"),
    ("hyundai",     "Hyundai",     "hyundai.png"),
    ("kia",         "Kia",         "kia.png"),
    ("seat",        "SEAT",        "seat.png"),
    ("skoda",       "Skoda",       "skoda.png"),
    ("porsche",     "Porsche",     "porsche.png"),
    ("ferrari",     "Ferrari",     "ferrari.png"),
    ("lamborghini", "Lamborghini", "lamborghini.png"),
    ("maserati",    "Maserati",    "maserati.png"),
    ("alfa_romeo",  "Alfa Romeo",  "alfa_romeo.png"),
    ("fiat",        "Fiat",        "fiat.png"),
    ("volvo",       "Volvo",       "volvo.png"),
    ("saab",        "Saab",        "saab.png"),
    ("land_rover",  "Land Rover",  "land_rover.png"),
    ("jaguar",      "Jaguar",      "jaguar.png"),
    ("mini",        "MINI",        "mini.png"),
    ("citroen",     "Citroën",     "citroen.png"),
    ("mitsubishi",  "Mitsubishi",  "mitsubishi.png"),
    ("subaru",      "Subaru",      "subaru.png"),
    ("lexus",       "Lexus",       "lexus.png"),
    ("infiniti",    "Infiniti",    "infiniti.png"),
    ("acura",       "Acura",       "acura.png"),
    ("chevrolet",   "Chevrolet",   "chevrolet.png"),
    ("dodge",       "Dodge",       "dodge.png"),
    ("jeep",        "Jeep",        "jeep.png"),
    ("chrysler",    "Chrysler",    "chrysler.png"),
    ("cadillac",    "Cadillac",    "cadillac.png"),
    ("buick",       "Buick",       "buick.png"),
    ("lincoln",     "Lincoln",     "lincoln.png"),
    ("tesla",       "Tesla",       "tesla.png"),
    ("genesis",     "Genesis",     "genesis.png"),
    ("bentley",     "Bentley",     "bentley.png"),
    ("rolls_royce", "Rolls-Royce", "rolls_royce.png"),
    ("aston_martin","Aston Martin","aston_martin.png"),
    ("mclaren",     "McLaren",     "mclaren.png"),
    ("bugatti",     "Bugatti",     "bugatti.png"),
    ("suzuki",      "Suzuki",      "suzuki.png"),
    ("dacia",       "Dacia",       "dacia.png"),
    ("cupra",       "Cupra",       "cupra.png"),
]

for i, row in enumerate(brands, 2):
    for j, val in enumerate(row, 1):
        ws_brands.cell(row=i, column=j, value=val)

style_rows(ws_brands, len(brands))
ws_brands.column_dimensions["A"].width = 20
ws_brands.column_dimensions["B"].width = 22
ws_brands.column_dimensions["C"].width = 24
ws_brands.freeze_panes = "A2"

# ================================================================
# Sheet 2: ENGINES
# ================================================================
ws_eng = wb.create_sheet("Engines")
cols_eng = ["engine_id", "engine_name"]
style_header(ws_eng, cols_eng)

engines_data = [
    ("vag_20tdi_150",        "2.0 TDI 150hp (EA288)"),
    ("bmw_b47_190",          "2.0d B47 190hp"),
    ("mercedes_om654_200",   "2.0d OM654 200hp"),
]
for i, row in enumerate(engines_data, 2):
    for j, val in enumerate(row, 1):
        ws_eng.cell(row=i, column=j, value=val)

style_rows(ws_eng, len(engines_data))
ws_eng.column_dimensions["A"].width = 28
ws_eng.column_dimensions["B"].width = 32
ws_eng.freeze_panes = "A2"

# ================================================================
# Sheet 3: MODELS
# ================================================================
ws_mod = wb.create_sheet("Models")
cols_mod = ["model_id", "brand_id", "model_name", "engine_id", "year_from", "year_to", "photo_file"]
style_header(ws_mod, cols_mod)

models_data = [
    ("audi_a4_20tdi",    "audi",        "A4 2.0 TDI",    "vag_20tdi_150",       2016, 2023, "audi_a4_20tdi.jpg"),
    ("audi_a6_20tdi",    "audi",        "A6 2.0 TDI",    "vag_20tdi_150",       2018, 2024, "audi_a6_20tdi.jpg"),
    ("vw_passat_20tdi",  "volkswagen",  "Passat 2.0 TDI","vag_20tdi_150",       2015, 2023, "vw_passat_20tdi.jpg"),
    ("bmw_320d",         "bmw",         "320d G20",       "bmw_b47_190",         2019, 2024, "bmw_320d.jpg"),
    ("bmw_520d",         "bmw",         "520d G30",       "bmw_b47_190",         2017, 2024, "bmw_520d.jpg"),
    ("mercedes_c220d",   "mercedes",    "C 220d W206",   "mercedes_om654_200",  2021, "",   "mercedes_c220d.jpg"),
]
for i, row in enumerate(models_data, 2):
    for j, val in enumerate(row, 1):
        ws_mod.cell(row=i, column=j, value=val)

style_rows(ws_mod, len(models_data))
for col, w in zip("ABCDEFG", [22, 18, 24, 26, 12, 12, 28]):
    ws_mod.column_dimensions[col].width = w
ws_mod.freeze_panes = "A2"

# ================================================================
# Sheet 4: STAGES
# ================================================================
ws_stg = wb.create_sheet("Stages")
cols_stg = ["engine_id", "stage", "orig_hp", "orig_torque", "tuned_hp", "tuned_torque"]
style_header(ws_stg, cols_stg)

stages_data = [
    ("vag_20tdi_150",       1, 150, 340, 185, 400),
    ("vag_20tdi_150",       2, 150, 340, 210, 440),
    ("vag_20tdi_150",       3, "",  "",  "",  ""),
    ("bmw_b47_190",         1, 190, 400, 230, 480),
    ("bmw_b47_190",         2, 190, 400, 260, 530),
    ("bmw_b47_190",         3, "",  "",  "",  ""),
    ("mercedes_om654_200",  1, 200, 440, 240, 500),
    ("mercedes_om654_200",  2, "",  "",  "",  ""),
    ("mercedes_om654_200",  3, "",  "",  "",  ""),
]
for i, row in enumerate(stages_data, 2):
    for j, val in enumerate(row, 1):
        ws_stg.cell(row=i, column=j, value=val)

style_rows(ws_stg, len(stages_data))
for col, w in zip("ABCDEF", [28, 8, 10, 14, 10, 14]):
    ws_stg.column_dimensions[col].width = w
ws_stg.freeze_panes = "A2"

# ================================================================
# Save
# ================================================================
wb.save("/Users/ronavaida/powercode-performance/cars.xlsx")
print("✅  cars.xlsx created successfully!")
