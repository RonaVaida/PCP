#!/usr/bin/env python3
"""
generate_json.py — Power Code Performance
==========================================
Reads cars.xlsx and writes data.json for the website.

Usage:
    python3 generate_json.py

Requirements:
    pip install openpyxl

Excel sheets expected:
  - Brands  : brand_id | brand_name | logo_file
  - Engines : engine_id | engine_name
  - Models  : model_id | brand_id | model_name | engine_id | year_from | year_to | photo_file
  - Stages  : engine_id | stage | orig_hp | orig_torque | tuned_hp | tuned_torque
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)

EXCEL_FILE = "cars.xlsx"
JSON_FILE  = "data.json"


def sheet_to_dicts(ws):
    """Convert a worksheet to a list of dicts using the first row as keys."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip empty rows
        d = {}
        for h, v in zip(headers, row):
            if h is None:
                continue
            # Normalise: strip whitespace, convert numbers to int where possible
            if isinstance(v, float) and v == int(v):
                v = int(v)
            elif isinstance(v, str):
                v = v.strip()
            d[str(h).strip()] = v
        rows.append(d)
    return rows


def main():
    xl_path = Path(EXCEL_FILE)
    if not xl_path.exists():
        print(f"ERROR: {EXCEL_FILE} not found in this folder.")
        sys.exit(1)

    wb = openpyxl.load_workbook(xl_path, data_only=True)

    required_sheets = ["Brands", "Engines", "Models", "Stages"]
    for s in required_sheets:
        if s not in wb.sheetnames:
            print(f"ERROR: Sheet '{s}' not found in {EXCEL_FILE}.")
            print(f"  Found sheets: {wb.sheetnames}")
            sys.exit(1)

    # ---- BRANDS ----
    brands = sheet_to_dicts(wb["Brands"])
    print(f"  Brands   : {len(brands)} rows")

    # ---- ENGINES ----
    engine_rows = sheet_to_dicts(wb["Engines"])
    engines = {e["engine_id"]: e for e in engine_rows}
    print(f"  Engines  : {len(engines)} rows")

    # ---- MODELS ----
    models = sheet_to_dicts(wb["Models"])
    # Convert year fields
    for m in models:
        for f in ("year_from", "year_to"):
            val = m.get(f)
            if val == "" or val is None:
                m[f] = None
            else:
                try:
                    m[f] = int(val)
                except (ValueError, TypeError):
                    m[f] = None
        # photo_file: empty string → None
        if not m.get("photo_file"):
            m["photo_file"] = None
    print(f"  Models   : {len(models)} rows")

    # ---- STAGES ----
    stage_rows = sheet_to_dicts(wb["Stages"])
    stages = {}
    for row in stage_rows:
        eid   = row.get("engine_id")
        stage = str(row.get("stage", "")).strip()
        if not eid or not stage:
            continue
        if eid not in stages:
            stages[eid] = {}
        # Only store if at least tuned_hp is filled
        tuned_hp = row.get("tuned_hp")
        if tuned_hp:
            stages[eid][stage] = {
                "orig_hp":      row.get("orig_hp"),
                "orig_torque":  row.get("orig_torque"),
                "tuned_hp":     row.get("tuned_hp"),
                "tuned_torque": row.get("tuned_torque"),
            }
        else:
            stages[eid][stage] = None
    print(f"  Stages   : {len(stage_rows)} rows")

    # ---- BUILD OUTPUT ----
    output = {
        "brands":  brands,
        "models":  models,
        "engines": engines,
        "stages":  stages,
    }

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅  Written to {JSON_FILE}  ({len(brands)} brands, {len(models)} models)")


if __name__ == "__main__":
    main()
