#!/usr/bin/env python3
"""
Populates cars.xlsx with all brands and models collected from br-performance.be
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- Styles ----
header_font  = Font(bold=True, color="FFFFFF", size=11)
red_fill     = PatternFill("solid", fgColor="E8001C")
alt_fill     = PatternFill("solid", fgColor="F2F2F2")
center_align = Alignment(horizontal="center", vertical="center")
wrap_align   = Alignment(wrap_text=True, vertical="center")
thin_border  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def style_header(ws, cols):
    ws.row_dimensions[1].height = 22
    for col, title in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = red_fill
        c.alignment = center_align
        c.border = thin_border

def style_rows(ws, num_rows):
    for row in range(2, num_rows + 2):
        fill = alt_fill if row % 2 == 0 else PatternFill()
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            if fill.fill_type:
                c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(vertical="center")

# ================================================================
# DATA
# ================================================================

BRANDS = [
    ("alfa_romeo",  "Alfa Romeo",   "alfa_romeo.png"),
    ("audi",        "Audi",         "audi.png"),
    ("bmw",         "BMW",          "bmw.png"),
    ("citroen",     "Citroën",      "citroen.png"),
    ("cupra",       "Cupra",        "cupra.png"),
    ("dacia",       "Dacia",        "dacia.png"),
    ("fiat",        "Fiat",         "fiat.png"),
    ("ford",        "Ford",         "ford.png"),
    ("honda",       "Honda",        "honda.png"),
    ("hyundai",     "Hyundai",      "hyundai.png"),
    ("infiniti",    "Infiniti",     "infiniti.png"),
    ("jaguar",      "Jaguar",       "jaguar.png"),
    ("kia",         "Kia",          "kia.png"),
    ("lamborghini", "Lamborghini",  "lamborghini.png"),
    ("landrover",   "Land Rover",   "landrover.png"),
    ("mazda",       "Mazda",        "mazda.png"),
    ("mercedes",    "Mercedes",     "mercedes.png"),
    ("mini",        "MINI",         "mini.png"),
    ("mitsubishi",  "Mitsubishi",   "mitsubishi.png"),
    ("nissan",      "Nissan",       "nissan.png"),
    ("opel",        "Opel",         "opel.png"),
    ("peugeot",     "Peugeot",      "peugeot.png"),
    ("porsche",     "Porsche",      "porsche.png"),
    ("renault",     "Renault",      "renault.png"),
    ("saab",        "Saab",         "saab.png"),
    ("seat",        "SEAT",         "seat.png"),
    ("skoda",       "Skoda",        "skoda.png"),
    ("subaru",      "Subaru",       "subaru.png"),
    ("suzuki",      "Suzuki",       "suzuki.png"),
    ("toyota",      "Toyota",       "toyota.png"),
    ("volkswagen",  "Volkswagen",   "volkswagen.png"),
    ("volvo",       "Volvo",        "volvo.png"),
]

# (brand_id, model_name)
MODELS_RAW = {
    "alfa_romeo": [
        "147","156","159","166","4C","Brennero","Brera","CrossWagon",
        "Giulia","Giulietta","GT","Junior","Milano","MiTo","Spider","Stelvio","Tonale"
    ],
    "audi": [
        "A1","A2","A3","A4","A4 Cabrio","A5","A6","A6 e-Tron","A7","A8",
        "e-Tron","e-Tron GT","Q2","Q3","Q4 e-Tron","Q5","Q6","Q6 e-Tron","Q7","Q8","Q8 E-Tron",
        "R8","RS3","RSQ3","RS4","RS5","RS6","RS7","RSQ8",
        "S1","S3","S4","S5","S6","S6 e-Tron","S7","S8","SQ2","SQ5","SQ6 e-Tron","SQ7","SQ8",
        "TT","TT S","TT RS"
    ],
    "bmw": [
        "i3","i4","i7","i8","iX","iX1","iX3",
        "Serie 1","Serie 2","Serie 2 GC","Serie 2 Gran/Active Tourer",
        "Serie 3","Serie 3 GT","Serie 4","Serie 4 GC",
        "Serie 5","Serie 5 GT","Serie 6","Serie 6 GC","Serie 6 GT",
        "Serie 7","Serie 8","Serie 8 GC",
        "1M Coupé","M2","M3","M4","M5","M6","M8",
        "X1","X2","X3","X3 M","X4","X4 M","X5","X5 M","X6","X6 M","X7","XM",
        "Z3","Z4","Z8"
    ],
    "citroen": [
        "Ami","Berlingo","C-Crosser","C-Elysée",
        "C1","C2","C3","C3 Aircross","C3 Picasso",
        "C4","C4 Aircross","C4 Cactus","C4 Picasso / Space Tourer","C4X",
        "C5 / C5 Aircross","C5X","C6","C8",
        "DS3","DS4","DS5","Jumper","Jumpy / SpaceTourer","Nemo","Picasso"
    ],
    "cupra": [
        "Ateca","Born","Formentor","Leon","Tavascan","Terramar"
    ],
    "dacia": [
        "Bigster","Dokker","Duster","Jogger","Lodgy","Logan","Sandero","Spring"
    ],
    "fiat": [
        "124 Spider","500 / 595 / 695","500L","500X","600","Brava","Bravo","Croma",
        "Doblo","Ducato","Fiorino","Freemont","Grande Panda","Grande Punto",
        "Idea","Linea","Marea","Multipla","Panda","Punto","Punto EVO",
        "Qubo","Scudo","Sedici","Stilo","Talento","Tipo / Tipo Cross","Ulysse"
    ],
    "ford": [
        "B-Max","Bronco","C-Max","EcoSport","Edge","Escape","Explorer",
        "F150","F250","F350","Fiesta","Focus","Fusion","Galaxy","GT",
        "Ka / Ka+","Kuga","Maverick","Mondeo","Mustang","Mustang Mach-E","Mustang Shelby",
        "Puma","Ranger","S-Max","Tourneo","Tourneo Connect","Tourneo Custom",
        "Transit","Transit Connect","Transit Custom"
    ],
    "honda": [
        "Accord","Civic","CR-V","CR-Z","E","FR-V","HR-V","Jazz","NSX","Z-RV"
    ],
    "hyundai": [
        "Accent","Bayon","Elantra","Getz","Grandeur","H serie","Ioniq",
        "i10","i20","i30","i40","Inster","ix20","ix35","ix55",
        "Kona","Nexo","Santa Fe","Sonata","Staria","Tucson","Veloster"
    ],
    "infiniti": [
        "EX","FX","G37/M37","M","Q30","Q50","Q60","Q70","QX30","QX60","QX70"
    ],
    "jaguar": [
        "E-Pace","F-Pace","F-Type / S","I-Pace","S-Type","X-Type","XE","XF","XJ","XK coupé","XKR","XKR-S"
    ],
    "kia": [
        "Carens","Carnival","Cee'd / Pro Cee'd","Cerato","EV6",
        "Magentis","Niro","Optima","Picanto","Rio",
        "Sorento","Soul","Sportage","Stinger","Stonic","Venga","X-Ceed"
    ],
    "lamborghini": [
        "Aventador","Gallardo","Huracan","Murcielago","Urus"
    ],
    "landrover": [
        "Defender","Discovery","Discovery Sport","Evoque","Freelander",
        "Range Rover / Range Rover Sport","Velar"
    ],
    "mazda": [
        "BT-50","CX-3","CX-30","CX-5","CX-60","CX-7","CX-80",
        "Mazda 2","Mazda 3","Mazda 5","Mazda 6","MX5","MX-30","RX8"
    ],
    "mercedes": [
        "A / A Berline","AMG GT Coupé / Roadster","AMG GT 4-door Coupé",
        "B","C","Citan","CL","CLA","CLE","CLK","CLS",
        "E / E Coupé / All Terrain","EQA","EQB","EQC","EQE","EQS","EQT","EQV",
        "G","GL","GLA","GLB","GLC / GLC Coupé","GLC EQ","GLE / GLE Coupé","GLK","GLS",
        "Maybach","ML","R","S","SL","SLC","SLK","SLS","Sprinter","T","V","Viano","Vito","X"
    ],
    "mini": [
        "Aceman","One / One D","Clubman","Cooper","Cooper S","Cooper S D",
        "Countryman","Paceman","Roadster/Coupé"
    ],
    "mitsubishi": [
        "ASX","Carisma","Colt","Eclipse","EVO","Grandis",
        "L200","Lancer","Outlander","Pajero"
    ],
    "nissan": [
        "350Z","370Z","Ariya","Almera","Cube","GTR","Juke","Leaf",
        "Micra","Murano","Note","NP 300 - Navara","NV200","Pathfinder",
        "Patrol","Primera","Qashqai","Sentra","Tiida","Townstar","X-Trail"
    ],
    "opel": [
        "Adam","Agila","Antara","Astra","Cascada","Combo","Corsa",
        "Crossland","Crossland X","Frontera","Grandland","Grandland X","GT",
        "Insignia / Insignia Grand Sport","Meriva","Mokka","Movano",
        "Signum","Tigra","Vectra","Vivaro","Zafira","Zafira Life"
    ],
    "peugeot": [
        "107","108","206","207","208","2008",
        "306","307","308","3008","406","407","408","508","5008",
        "607","807","Bipper","Boxer","Expert / Traveller","Partner","RCZ","Rifter"
    ],
    "porsche": [
        "911","Boxster","Carrera GT","Cayenne","Cayman","Macan","Panamera","Taycan"
    ],
    "renault": [
        "Alaskan","Arkana","Austral","Captur","Clio","Espace","Fluence",
        "Kadjar","Kangoo","Koleos","Laguna","Latitude","Master","Megane",
        "Megane E-Tech","Modus","Rafale","R5","Scenic / Grand Scenic",
        "Symbioz","Talisman","Trafic","Twingo","ZOE"
    ],
    "saab": [
        "9-3","9-4X","9-5"
    ],
    "seat": [
        "Alhambra","Altea","Altea XL / Freetrack","Arona","Ateca",
        "Cordoba","Exeo","Ibiza","Leon","Mii","Tarraco","Toledo"
    ],
    "skoda": [
        "Citigo","Elroq","Enyaq","Fabia","Kamiq","Karoq","Kodiaq",
        "Kushaq","Octavia","Rapid","Roomster","Scala","Superb","Yeti"
    ],
    "subaru": [
        "BRZ","Forester","Impreza","Legacy","Levorg","Outback","Soltera","XV"
    ],
    "suzuki": [
        "Grand Vitara","Ignis","Baleno","Jimny","Liana","Splash",
        "Swace","Swift","SX-4","SX-4 S-Cross","Vitara"
    ],
    "toyota": [
        "Auris","Avensis","Aygo","Aygo X","BZ4X","Camry","C-HR","Corolla",
        "GR86","GT86","Hilux","IQ","Land Cruiser","Mirai","Prius / Prius+",
        "ProAce / ProAce Verso","Rav4","Supra (GR)","Urban Cruiser","Verso","Yaris","Yaris Cross"
    ],
    "volkswagen": [
        "Amarok","Arteon","Atlas / Teramont","Bora","Caddy","Caravelle",
        "Coccinelle / New Beetle","Crafter","Eos","Fox","Golf",
        "ID.3","ID.4","ID.5","ID.7","ID.Buzz",
        "Jetta","LT","Lupo","Multivan / California","Passat","Passat CC / CC",
        "Phaeton","Polo","Scirocco","Sharan","T-Cross","T-Roc","Taigo",
        "Tiguan","Touareg","Touran","Transporter","Up!"
    ],
    "volvo": [
        "C30","C40","C70","EX30","EX40","EX90",
        "P1800 Cyan","S40 / V50","V40 / V40 CC","S60 / V60",
        "V70","S80","S90 / V90",
        "XC 40","XC 60","XC 70","XC 90"
    ],
}

# ================================================================
# Sheet 1: BRANDS
# ================================================================
ws_brands = wb.active
ws_brands.title = "Brands"
cols_b = ["brand_id", "brand_name", "logo_file"]
style_header(ws_brands, cols_b)

for i, (bid, bname, blogo) in enumerate(BRANDS, 2):
    ws_brands.cell(row=i, column=1, value=bid)
    ws_brands.cell(row=i, column=2, value=bname)
    ws_brands.cell(row=i, column=3, value=blogo)

style_rows(ws_brands, len(BRANDS))
ws_brands.column_dimensions["A"].width = 20
ws_brands.column_dimensions["B"].width = 22
ws_brands.column_dimensions["C"].width = 26
ws_brands.freeze_panes = "A2"

# ================================================================
# Sheet 2: ENGINES  (sample — user fills in)
# ================================================================
ws_eng = wb.create_sheet("Engines")
cols_e = ["engine_id", "engine_name"]
style_header(ws_eng, cols_e)

engines_data = [
    ("vag_20tdi_150",       "2.0 TDI 150hp (EA288)"),
    ("bmw_b47_190",         "2.0d B47 190hp"),
    ("mercedes_om654_200",  "2.0d OM654 200hp"),
]
for i, row in enumerate(engines_data, 2):
    for j, val in enumerate(row, 1):
        ws_eng.cell(row=i, column=j, value=val)

style_rows(ws_eng, len(engines_data))
ws_eng.column_dimensions["A"].width = 28
ws_eng.column_dimensions["B"].width = 36
ws_eng.freeze_panes = "A2"

# ================================================================
# Sheet 3: MODELS
# ================================================================
ws_mod = wb.create_sheet("Models")
cols_m = ["model_id", "brand_id", "model_name", "engine_id", "year_from", "year_to", "photo_file"]
style_header(ws_mod, cols_m)

row_num = 2
model_rows = []
for brand_id, model_name in [
    (bid, m) for bid, models in MODELS_RAW.items() for m in models
]:
    # auto-generate model_id: brand_id + sanitised model name
    slug = model_name.lower()
    for ch in [" ", "/", ".", "-", "(", ")", "+"]:
        slug = slug.replace(ch, "_")
    # collapse multiple underscores
    while "__" in slug:
        slug = slug.replace("__", "_")
    slug = slug.strip("_")
    model_id = f"{brand_id}_{slug}"
    model_rows.append((model_id, brand_id, model_name, "", "", "", ""))

for i, row in enumerate(model_rows, 2):
    for j, val in enumerate(row, 1):
        ws_mod.cell(row=i, column=j, value=val)

style_rows(ws_mod, len(model_rows))
for col, w in zip("ABCDEFG", [40, 18, 30, 26, 12, 12, 30]):
    ws_mod.column_dimensions[col].width = w
ws_mod.freeze_panes = "A2"

# ================================================================
# Sheet 4: STAGES  (empty template, user fills)
# ================================================================
ws_stg = wb.create_sheet("Stages")
cols_s = ["engine_id", "stage", "orig_hp", "orig_torque", "tuned_hp", "tuned_torque"]
style_header(ws_stg, cols_s)

stages_sample = [
    ("vag_20tdi_150",       1, 150, 340, 185, 400),
    ("vag_20tdi_150",       2, 150, 340, 210, 440),
    ("vag_20tdi_150",       3, "",  "",  "",  ""),
    ("bmw_b47_190",         1, 190, 400, 230, 480),
    ("bmw_b47_190",         2, 190, 400, 260, 530),
    ("bmw_b47_190",         3, "",  "",  "",  ""),
    ("mercedes_om654_200",  1, 200, 440, 240, 500),
    ("mercedes_om654_200",  2, "",  "",  "",  ""),
    ("mercedes_om654_200",  3, "",  "",  "",  ""),
]
for i, row in enumerate(stages_sample, 2):
    for j, val in enumerate(row, 1):
        ws_stg.cell(row=i, column=j, value=val)

style_rows(ws_stg, len(stages_sample))
for col, w in zip("ABCDEF", [28, 8, 10, 14, 10, 14]):
    ws_stg.column_dimensions[col].width = w
ws_stg.freeze_panes = "A2"

# ================================================================
# Save
# ================================================================
out_path = "/Users/ronavaida/powercode-performance/cars.xlsx"
wb.save(out_path)

total_brands = len(BRANDS)
total_models = len(model_rows)
print(f"✅  cars.xlsx saved!")
print(f"   Brands : {total_brands}")
print(f"   Models : {total_models}")
print(f"   Path   : {out_path}")
